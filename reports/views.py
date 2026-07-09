from django import forms
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from .models import PreacherPersonalInfo, TeamLeaderPersonalData, CongregationUpdate, KCMBIFieldReport, Zone
from .forms import (
    PreacherPersonalInfoForm, TeamLeaderPersonalDataForm,
    CongregationUpdateForm, KCMBIFieldReportForm, ZoneForm
)
from .utils.pdf_generator import (
    generate_preacher_pdf, generate_teamleader_pdf,
    generate_congregation_pdf, generate_fieldreport_pdf,
    generate_zone_pdf
)


def get_role(user):
    profile = getattr(user, "profile", None)
    return profile.role if profile else "admin"


def get_preacher_queryset(user):
    role = get_role(user)
    if role == "preacher":
        preacher = getattr(user, "preacher_profile", None)
        if preacher:
            return PreacherPersonalInfo.objects.filter(id=preacher.id)
        return PreacherPersonalInfo.objects.none()
    elif role == "team_leader":
        leader = getattr(user, "teamleader_profile", None)
        if leader:
            return PreacherPersonalInfo.objects.filter(team_leader_ref=leader)
        return PreacherPersonalInfo.objects.none()
    return PreacherPersonalInfo.objects.all()


def get_congregation_queryset(user):
    role = get_role(user)
    if role == "preacher":
        preacher = getattr(user, "preacher_profile", None)
        if preacher:
            return CongregationUpdate.objects.filter(preacher_ref=preacher)
        return CongregationUpdate.objects.none()
    elif role == "team_leader":
        leader = getattr(user, "teamleader_profile", None)
        if leader:
            zone_ids = Zone.objects.filter(team_leader=leader).values_list("id", flat=True)
            return CongregationUpdate.objects.filter(zone__id__in=zone_ids)
        return CongregationUpdate.objects.none()
    return CongregationUpdate.objects.all()


def get_fieldreport_queryset(user):
    role = get_role(user)
    if role == "team_leader":
        leader = getattr(user, "teamleader_profile", None)
        if leader:
            return KCMBIFieldReport.objects.filter(team_leader_ref=leader)
        return KCMBIFieldReport.objects.none()
    elif role == "preacher":
        return KCMBIFieldReport.objects.none()
    return KCMBIFieldReport.objects.all()


def get_teamleader_queryset(user):
    role = get_role(user)
    if role == "admin":
        return TeamLeaderPersonalData.objects.all()
    elif role == "team_leader":
        leader = getattr(user, "teamleader_profile", None)
        if leader:
            return TeamLeaderPersonalData.objects.filter(id=leader.id)
        return TeamLeaderPersonalData.objects.none()
    return TeamLeaderPersonalData.objects.none()


def get_zone_queryset(user):
    role = get_role(user)
    if role == "admin":
        return Zone.objects.all()
    elif role == "team_leader":
        leader = getattr(user, "teamleader_profile", None)
        if leader:
            return Zone.objects.filter(team_leader=leader)
        return Zone.objects.none()
    return Zone.objects.none()


@login_required
def preacher_list(request):
    query = request.GET.get("q", "")
    records = get_preacher_queryset(request.user).order_by("-created_at")
    if query:
        records = records.filter(name_of_preacher__icontains=query)
    return render(request, "reports/preacher_list.html", {"records": records, "query": query, "role": get_role(request.user)})


@login_required
def preacher_add(request):
    role = get_role(request.user)
    if role == "preacher":
        messages.error(request, "Preachers cannot add new preacher records.")
        return redirect("preacher_list")
    if request.method == "POST":
        form = PreacherPersonalInfoForm(request.POST, request.FILES)
        if form.is_valid():
            obj = form.save(commit=False)
            if role == "team_leader":
                leader = getattr(request.user, "teamleader_profile", None)
                if leader:
                    obj.team_leader_ref = leader
            elif role == "admin" and form.cleaned_data.get("team_leader_ref"):
                obj.team_leader_ref = form.cleaned_data["team_leader_ref"]
            obj.save()
            messages.success(request, "Preacher record added successfully!")
            return redirect("preacher_list")
    else:
        form = PreacherPersonalInfoForm()
        if role == "admin":
            form.fields["team_leader_ref"] = forms.ModelChoiceField(
                queryset=TeamLeaderPersonalData.objects.all(),
                required=False, label="Team Leader",
                widget=forms.Select(attrs={"class": "form-control"})
            )
        elif role == "team_leader":
            leader = getattr(request.user, "teamleader_profile", None)
            if leader:
                form.fields["zone"].queryset = Zone.objects.filter(team_leader=leader)
    return render(request, "reports/preacher_form.html", {"form": form, "title": "Add Preacher Personal Info"})


@login_required
def approve_preacher(request, pk):
    if get_role(request.user) != "admin":
        messages.error(request, "Only admins can approve preachers.")
        return redirect("preacher_list")
    preacher = get_object_or_404(PreacherPersonalInfo, pk=pk)
    preacher.is_approved = True
    preacher.is_rejected = False
    preacher.save()
    messages.success(request, f'Preacher "{preacher.name_of_preacher}" approved!')
    return redirect("preacher_list")


@login_required
def reject_preacher(request, pk):
    if get_role(request.user) != "admin":
        messages.error(request, "Only admins can reject preachers.")
        return redirect("preacher_list")
    preacher = get_object_or_404(PreacherPersonalInfo, pk=pk)
    preacher.is_rejected = True
    preacher.is_approved = False
    preacher.save()
    messages.success(request, f'Preacher "{preacher.name_of_preacher}" rejected.')
    return redirect("preacher_list")


@login_required
def preacher_detail(request, pk):
    record = get_object_or_404(PreacherPersonalInfo, pk=pk)
    return render(request, "reports/preacher_detail.html", {"record": record})


@login_required
def preacher_edit(request, pk):
    role = get_role(request.user)
    record = get_object_or_404(PreacherPersonalInfo, pk=pk)
    if role == "team_leader":
        leader = getattr(request.user, "teamleader_profile", None)
        if not leader or record.team_leader_ref != leader:
            messages.error(request, "You can only edit preachers in your team.")
            return redirect("preacher_list")
    if request.method == "POST":
        form = PreacherPersonalInfoForm(request.POST, request.FILES, instance=record)
        if form.is_valid():
            form.save()
            messages.success(request, "Preacher record updated successfully!")
            return redirect("preacher_list")
    else:
        form = PreacherPersonalInfoForm(instance=record)
        if role == "team_leader":
            leader = getattr(request.user, "teamleader_profile", None)
            if leader and record.team_leader_ref == leader:
                form.fields["zone"].queryset = Zone.objects.filter(team_leader=leader)
    return render(request, "reports/preacher_form.html", {"form": form, "title": "Edit Preacher Personal Info"})


@login_required
def preacher_delete(request, pk):
    record = get_object_or_404(PreacherPersonalInfo, pk=pk)
    if request.method == "POST":
        record.delete()
        messages.success(request, "Preacher record deleted successfully!")
        return redirect("preacher_list")
    return render(request, "reports/preacher_confirm_delete.html", {"record": record})


@login_required
def preacher_pdf(request, pk):
    record = get_object_or_404(PreacherPersonalInfo, pk=pk)
    buffer = generate_preacher_pdf(record)
    filename = "preacher_" + record.name_of_preacher.replace(" ", "_") + ".pdf"
    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="' + filename + '"'
    return response


@login_required
def teamleader_list(request):
    query = request.GET.get("q", "")
    records = get_teamleader_queryset(request.user).order_by("-created_at")
    if query:
        records = records.filter(name__icontains=query)
    return render(request, "reports/teamleader_list.html", {"records": records, "query": query, "role": get_role(request.user)})


@login_required
def teamleader_add(request):
    if get_role(request.user) != "admin":
        messages.error(request, "Only admins can add team leaders.")
        return redirect("teamleader_list")
    if request.method == "POST":
        form = TeamLeaderPersonalDataForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Team Leader record added successfully!")
            return redirect("teamleader_list")
    else:
        form = TeamLeaderPersonalDataForm()
    return render(request, "reports/teamleader_form.html", {"form": form, "title": "Add Team Leader Personal Data"})


@login_required
def teamleader_detail(request, pk):
    record = get_object_or_404(TeamLeaderPersonalData, pk=pk)
    return render(request, "reports/teamleader_detail.html", {"record": record})


@login_required
def teamleader_edit(request, pk):
    record = get_object_or_404(TeamLeaderPersonalData, pk=pk)
    if request.method == "POST":
        form = TeamLeaderPersonalDataForm(request.POST, request.FILES, instance=record)
        if form.is_valid():
            form.save()
            messages.success(request, "Team Leader record updated successfully!")
            return redirect("teamleader_list")
    else:
        form = TeamLeaderPersonalDataForm(instance=record)
    return render(request, "reports/teamleader_form.html", {"form": form, "title": "Edit Team Leader Personal Data"})


@login_required
def teamleader_delete(request, pk):
    record = get_object_or_404(TeamLeaderPersonalData, pk=pk)
    if request.method == "POST":
        record.delete()
        messages.success(request, "Team Leader record deleted successfully!")
        return redirect("teamleader_list")
    return render(request, "reports/teamleader_confirm_delete.html", {"record": record})


@login_required
def teamleader_pdf(request, pk):
    record = get_object_or_404(TeamLeaderPersonalData, pk=pk)
    buffer = generate_teamleader_pdf(record)
    filename = "teamleader_" + record.name.replace(" ", "_") + ".pdf"
    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="' + filename + '"'
    return response


@login_required
def export_teamleaders_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    records = get_teamleader_queryset(request.user).order_by("-created_at")
    wb = Workbook()
    ws = wb.active
    ws.title = "Team Leaders"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["Name", "Email", "Cell Number", "Date of Birth", "Address",
                "Bank Name", "Account Number", "IFSC Code", "KCMBIs Submitting"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row, r in enumerate(records, 2):
        vals = [r.name, r.email_address, r.cell_number, str(r.date_of_birth),
                r.address, r.bank_name, r.account_number, r.ifsc_code, r.number_of_kcmbis_submitting]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = thin_border

    widths = [25, 30, 18, 15, 40, 22, 22, 18, 18]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col)].width = w

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="team_leaders.xlsx"'
    wb.save(response)
    return response


@login_required
def congregation_list(request):
    query = request.GET.get("q", "")
    records = get_congregation_queryset(request.user).order_by("-created_at")
    if query:
        records = records.filter(name_of_preacher__icontains=query)
    return render(request, "reports/congregation_list.html", {"records": records, "query": query, "role": get_role(request.user)})


@login_required
def congregation_add(request):
    role = get_role(request.user)
    if request.method == "POST":
        form = CongregationUpdateForm(request.POST, request.FILES)
        if form.is_valid():
            obj = form.save(commit=False)
            if role == "preacher":
                preacher = getattr(request.user, "preacher_profile", None)
                if preacher:
                    obj.name_of_preacher = preacher.name_of_preacher
                    obj.preacher_ref = preacher
                    if preacher.zone:
                        obj.zone = preacher.zone
            obj.save()
            messages.success(request, "Congregation update added successfully!")
            return redirect("congregation_list")
    else:
        form = CongregationUpdateForm()
        if role == "preacher":
            preacher = getattr(request.user, "preacher_profile", None)
            if preacher:
                form.fields["name_of_preacher"].initial = preacher.name_of_preacher
                form.fields["name_of_preacher"].widget.attrs["readonly"] = True
                if preacher.zone:
                    form.fields["zone"].initial = preacher.zone
                    form.fields["zone"].queryset = Zone.objects.filter(pk=preacher.zone.pk)
                    form.fields["zone"].widget.attrs["disabled"] = True
        elif role == "team_leader":
            leader = getattr(request.user, "teamleader_profile", None)
            if leader:
                form.fields["zone"].queryset = Zone.objects.filter(team_leader=leader)
    return render(request, "reports/congregation_form.html", {"form": form, "title": "Add Congregation Update"})


@login_required
def congregation_detail(request, pk):
    role = get_role(request.user)
    record = get_object_or_404(CongregationUpdate, pk=pk)
    if role == "preacher":
        preacher = getattr(request.user, "preacher_profile", None)
        if not preacher or record.preacher_ref != preacher:
            messages.error(request, "You can only view your own congregation updates.")
            return redirect("congregation_list")
    elif role == "team_leader":
        leader = getattr(request.user, "teamleader_profile", None)
        if leader and record.zone and record.zone.team_leader != leader:
            messages.error(request, "You can only view congregations in your zones.")
            return redirect("congregation_list")
    return render(request, "reports/congregation_detail.html", {"record": record})


@login_required
def congregation_edit(request, pk):
    role = get_role(request.user)
    record = get_object_or_404(CongregationUpdate, pk=pk)
    if role == "preacher":
        preacher = getattr(request.user, "preacher_profile", None)
        if not preacher or record.preacher_ref != preacher:
            messages.error(request, "You can only edit your own congregation updates.")
            return redirect("congregation_list")
    elif role == "team_leader":
        leader = getattr(request.user, "teamleader_profile", None)
        if leader and record.zone and record.zone.team_leader != leader:
            messages.error(request, "You can only edit congregations in your zones.")
            return redirect("congregation_list")
    if request.method == "POST":
        form = CongregationUpdateForm(request.POST, request.FILES, instance=record)
        if form.is_valid():
            obj = form.save(commit=False)
            if role == "preacher":
                preacher = getattr(request.user, "preacher_profile", None)
                if preacher and preacher.zone:
                    obj.zone = preacher.zone
            obj.save()
            messages.success(request, "Congregation update edited successfully!")
            return redirect("congregation_list")
    else:
        form = CongregationUpdateForm(instance=record)
        if role == "preacher":
            form.fields["name_of_preacher"].widget.attrs["readonly"] = True
            preacher = getattr(request.user, "preacher_profile", None)
            if preacher and preacher.zone:
                form.fields["zone"].queryset = Zone.objects.filter(pk=preacher.zone.pk)
                form.fields["zone"].widget.attrs["disabled"] = True
        elif role == "team_leader":
            leader = getattr(request.user, "teamleader_profile", None)
            if leader:
                form.fields["zone"].queryset = Zone.objects.filter(team_leader=leader)
    return render(request, "reports/congregation_form.html", {"form": form, "title": "Edit Congregation Update"})


@login_required
def congregation_delete(request, pk):
    role = get_role(request.user)
    record = get_object_or_404(CongregationUpdate, pk=pk)
    if role == "preacher":
        preacher = getattr(request.user, "preacher_profile", None)
        if not preacher or record.preacher_ref != preacher:
            messages.error(request, "You can only delete your own congregation updates.")
            return redirect("congregation_list")
    if request.method == "POST":
        record.delete()
        messages.success(request, "Congregation update deleted successfully!")
        return redirect("congregation_list")
    return render(request, "reports/congregation_confirm_delete.html", {"record": record})


@login_required
def export_congregations_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    records = get_congregation_queryset(request.user).order_by("-created_at")
    wb = Workbook()
    ws = wb.active
    ws.title = "Congregations"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["Preacher", "Team Leader", "Zone", "Village", "Month", "Bible Studies", "Baptisms",
               "Members", "Benevolent Aid", "Weekly Giving"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row, r in enumerate(records, 2):
        zone_text = f"{r.zone.zone_number} {r.zone.zone_name}" if r.zone else ""
        tl_name = r.zone.team_leader.name if r.zone and r.zone.team_leader else ""
        vals = [r.name_of_preacher, tl_name, zone_text, r.name_of_village, r.month_of_reporting,
                r.bible_studies_meetings_count, r.baptisms_count, r.church_members_count,
                str(r.benevolent_aid_received), str(r.average_weekly_giving)]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = thin_border

    widths = [25, 22, 20, 22, 16, 14, 12, 12, 16, 16]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col)].width = w

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="congregations.xlsx"'
    wb.save(response)
    return response


@login_required
def congregation_pdf(request, pk):
    record = get_object_or_404(CongregationUpdate, pk=pk)
    buffer = generate_congregation_pdf(record)
    filename = "congregation_" + record.name_of_preacher.replace(" ", "_") + ".pdf"
    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="' + filename + '"'
    return response


@login_required
def fieldreport_list(request):
    query = request.GET.get("q", "")
    records = get_fieldreport_queryset(request.user).order_by("-created_at")
    if query:
        records = records.filter(team_leader__icontains=query)
    return render(request, "reports/fieldreport_list.html", {"records": records, "query": query, "role": get_role(request.user)})


@login_required
def fieldreport_add(request):
    role = get_role(request.user)
    if role == "preacher":
        messages.error(request, "Preachers cannot add field reports.")
        return redirect("fieldreport_list")
    leader = getattr(request.user, "teamleader_profile", None) if role == "team_leader" else None
    tl_kwargs = {"team_leader_obj": leader} if leader else {}
    if request.method == "POST":
        form = KCMBIFieldReportForm(request.POST, request.FILES, **tl_kwargs)
        if form.is_valid():
            obj = form.save(commit=False)
            if role == "team_leader" and leader:
                obj.team_leader_ref = leader
                obj.team_leader = leader.name
            obj.save()
            messages.success(request, "Field report added successfully!")
            return redirect("fieldreport_list")
    else:
        form = KCMBIFieldReportForm(**tl_kwargs)
        if role == "team_leader" and leader:
            form.fields["team_leader"].initial = leader.name
            form.fields["team_leader"].widget.attrs["readonly"] = True
            preachers = PreacherPersonalInfo.objects.filter(team_leader_ref=leader).values_list('name_of_preacher', flat=True)
            if preachers:
                form.fields["preachers_in_attendance"].initial = '\n'.join(preachers)
                form.fields["preachers_in_attendance"].widget.attrs["readonly"] = True
    return render(request, "reports/fieldreport_form.html", {"form": form, "title": "Add KCMBI Field Report"})


@login_required
def fieldreport_detail(request, pk):
    record = get_object_or_404(KCMBIFieldReport, pk=pk)
    return render(request, "reports/fieldreport_detail.html", {"record": record})


@login_required
def fieldreport_edit(request, pk):
    role = get_role(request.user)
    record = get_object_or_404(KCMBIFieldReport, pk=pk)
    leader = record.team_leader_ref or (getattr(request.user, "teamleader_profile", None) if role == "team_leader" else None)
    tl_kwargs = {"team_leader_obj": leader} if leader else {}
    if request.method == "POST":
        form = KCMBIFieldReportForm(request.POST, request.FILES, instance=record, **tl_kwargs)
        if form.is_valid():
            obj = form.save(commit=False)
            if role == "team_leader" and leader:
                obj.team_leader_ref = leader
                obj.team_leader = leader.name
            obj.save()
            messages.success(request, "Field report updated successfully!")
            return redirect("fieldreport_list")
    else:
        form = KCMBIFieldReportForm(instance=record, **tl_kwargs)
        if role == "team_leader":
            form.fields["team_leader"].widget.attrs["readonly"] = True
    return render(request, "reports/fieldreport_form.html", {"form": form, "title": "Edit KCMBI Field Report"})


@login_required
def fieldreport_delete(request, pk):
    record = get_object_or_404(KCMBIFieldReport, pk=pk)
    if request.method == "POST":
        record.delete()
        messages.success(request, "Field report deleted successfully!")
        return redirect("fieldreport_list")
    return render(request, "reports/fieldreport_confirm_delete.html", {"record": record})


@login_required
def fieldreport_pdf(request, pk):
    record = get_object_or_404(KCMBIFieldReport, pk=pk)
    buffer = generate_fieldreport_pdf(record)
    filename = "fieldreport_" + record.team_leader.replace(" ", "_") + ".pdf"
    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="' + filename + '"'
    return response


@login_required
def export_fieldreports_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    records = get_fieldreport_queryset(request.user).order_by("-created_at")
    wb = Workbook()
    ws = wb.active
    ws.title = "Field Reports"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["Team Leader", "Zone", "Meeting Date", "KCMBI Number", "Topic/Text",
               "Meeting Address", "Preachers", "Group Concerns"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row, r in enumerate(records, 2):
        zone_text = f"{r.zone.zone_number} {r.zone.zone_name}" if r.zone else ""
        vals = [r.team_leader, zone_text, str(r.meeting_date_time), r.kcmbi_number,
                r.class_topic_or_text, r.meeting_address, r.preachers_in_attendance, r.group_concerns]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = thin_border

    widths = [22, 18, 20, 16, 28, 35, 35, 30]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col)].width = w

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="field_reports.xlsx"'
    wb.save(response)
    return response


@login_required
def zone_list(request):
    role = get_role(request.user)
    if role == "preacher":
        messages.error(request, "Preachers cannot access zones.")
        return redirect("dashboard")
    query = request.GET.get("q", "")
    records = get_zone_queryset(request.user).order_by("zone_number")
    if query:
        records = records.filter(zone_number__icontains=query) | records.filter(zone_name__icontains=query)
    return render(request, "reports/zone_list.html", {"records": records, "query": query, "role": get_role(request.user)})


@login_required
def zone_add(request):
    role = get_role(request.user)
    if role == "preacher":
        messages.error(request, "Preachers cannot manage zones.")
        return redirect("dashboard")
    if request.method == "POST":
        form = ZoneForm(request.POST)
        if form.is_valid():
            zone = form.save(commit=False)
            if role == "team_leader" and not zone.team_leader_id:
                leader = getattr(request.user, "teamleader_profile", None)
                if leader:
                    zone.team_leader = leader
            zone.save()
            messages.success(request, f"Zone {zone.zone_number} created successfully!")
            return redirect("zone_list")
    else:
        form = ZoneForm()
        if role == "team_leader":
            leader = getattr(request.user, "teamleader_profile", None)
            if leader:
                form.fields["team_leader"].queryset = TeamLeaderPersonalData.objects.filter(id=leader.id)
                form.fields["team_leader"].initial = leader
    return render(request, "reports/zone_form.html", {"form": form, "title": "Add Zone"})


@login_required
def zone_detail(request, pk):
    role = get_role(request.user)
    record = get_object_or_404(Zone, pk=pk)
    if role == "preacher":
        preacher = getattr(request.user, "preacher_profile", None)
        if not preacher or preacher.zone != record:
            messages.error(request, "You can only view your own zone.")
            return redirect("dashboard")
    elif role == "team_leader":
        leader = getattr(request.user, "teamleader_profile", None)
        if not leader or record.team_leader != leader:
            messages.error(request, "You can only view your own zones.")
            return redirect("dashboard")
    preachers = record.preachers.all()
    congregations = record.congregation_updates.all()
    return render(request, "reports/zone_detail.html", {
        "record": record, "preachers": preachers, "congregations": congregations,
        "role": role,
    })


@login_required
def zone_edit(request, pk):
    role = get_role(request.user)
    record = get_object_or_404(Zone, pk=pk)
    if role == "team_leader":
        leader = getattr(request.user, "teamleader_profile", None)
        if not leader or record.team_leader != leader:
            messages.error(request, "You can only edit your own zones.")
            return redirect("zone_list")
    if request.method == "POST":
        form = ZoneForm(request.POST, instance=record)
        if form.is_valid():
            form.save()
            messages.success(request, "Zone updated successfully!")
            return redirect("zone_list")
    else:
        form = ZoneForm(instance=record)
    return render(request, "reports/zone_form.html", {"form": form, "title": "Edit Zone"})


@login_required
def zone_delete(request, pk):
    role = get_role(request.user)
    record = get_object_or_404(Zone, pk=pk)
    if role == "team_leader":
        leader = getattr(request.user, "teamleader_profile", None)
        if not leader or record.team_leader != leader:
            messages.error(request, "You can only delete your own zones.")
            return redirect("zone_list")
    if request.method == "POST":
        record.delete()
        messages.success(request, "Zone deleted successfully!")
        return redirect("zone_list")
    return render(request, "reports/zone_confirm_delete.html", {"record": record})


@login_required
def zone_pdf(request, pk):
    record = get_object_or_404(Zone, pk=pk)
    buffer = generate_zone_pdf(record)
    filename = f"zone_{record.zone_number}.pdf"
    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="' + filename + '"'
    return response


@login_required
def zone_preachers_excel(request):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    role = get_role(request.user)
    if role == "preacher":
        messages.error(request, "Preachers cannot access this.")
        return redirect("dashboard")

    if role == "team_leader":
        leader = getattr(request.user, "teamleader_profile", None)
        zones = Zone.objects.filter(team_leader=leader).prefetch_related("preachers").order_by("zone_number") if leader else Zone.objects.none()
    else:
        zones = Zone.objects.all().prefetch_related("preachers").order_by("zone_number")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Zone Wise Preachers"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    headers = ["Zone #", "Zone Name", "Team Leader", "Preacher Name", "Cell Number", "Congregation Address", "Bank Name", "Account Number", "IFSC Code"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    row = 2
    for z in zones:
        preachers = z.preachers.all()
        if preachers:
            for p in preachers:
                ws.cell(row=row, column=1, value=z.zone_number)
                ws.cell(row=row, column=2, value=z.zone_name or "")
                ws.cell(row=row, column=3, value=z.team_leader.name if z.team_leader else "")
                ws.cell(row=row, column=4, value=p.name_of_preacher)
                ws.cell(row=row, column=5, value=p.cell_number)
                ws.cell(row=row, column=6, value=p.congregation_address)
                ws.cell(row=row, column=7, value=p.bank_name)
                ws.cell(row=row, column=8, value=p.account_number)
                ws.cell(row=row, column=9, value=p.ifsc_code)
                row += 1
        else:
            ws.cell(row=row, column=1, value=z.zone_number)
            ws.cell(row=row, column=2, value=z.zone_name or "")
            ws.cell(row=row, column=3, value=z.team_leader.name if z.team_leader else "")
            ws.cell(row=row, column=4, value="No preachers")
            row += 1

    widths = [12, 18, 22, 25, 18, 30, 20, 22, 18]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col)].width = w

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="zone_wise_preachers.xlsx"'
    wb.save(response)
    return response


@login_required
def export_zones_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    records = get_zone_queryset(request.user).order_by("zone_number")
    wb = Workbook()
    ws = wb.active
    ws.title = "Zones"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["Zone Number", "Zone Name", "Team Leader", "Location", "Description", "Preachers"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row, r in enumerate(records, 2):
        vals = [r.zone_number, r.zone_name or "",
                r.team_leader.name if r.team_leader else "",
                r.location or "", r.description or "",
                r.preachers.count()]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = thin_border

    widths = [14, 20, 25, 20, 35, 12]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col)].width = w

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="zones.xlsx"'
    wb.save(response)
    return response


@login_required
def autocomplete_preachers(request):
    term = request.GET.get("term", "")
    qs = PreacherPersonalInfo.objects.filter(name_of_preacher__icontains=term)
    role = get_role(request.user)
    if role == "team_leader":
        leader = getattr(request.user, "teamleader_profile", None)
        if leader:
            qs = qs.filter(team_leader_ref=leader)
    preachers = qs.select_related("zone", "team_leader_ref").values(
        "id", "name_of_preacher", "zone__id", "zone__zone_number", "zone__zone_name", "team_leader_ref__name"
    )[:20]
    return JsonResponse(list(preachers), safe=False)
